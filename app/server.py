#!/usr/bin/env python3
"""
Badge Scanner Server
Serves the badge scanner web app and handles AI identification via Claude vision.
"""

import json
import os
import base64
import uuid
import time
import io
import subprocess
import sys
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import urlparse
import urllib.request
import urllib.error

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

try:
    from PIL import Image
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'Pillow', '-q'])
    from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PORT = 8282
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PHOTOS_DIR = os.path.join(BASE_DIR, "photos")
COLLECTION_FILE = os.path.join(BASE_DIR, "..", "state", "badge-collection.json")
OPENCLAW_CONFIG = os.path.expanduser("~/.openclaw/openclaw.json")

os.makedirs(PHOTOS_DIR, exist_ok=True)


def remove_background(raw_bytes):
    try:
        from rembg import remove as rembg_remove
        result = rembg_remove(raw_bytes)
        # convert PNG result back to JPEG on white bg
        img = Image.open(io.BytesIO(result)).convert("RGBA")
        bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
        bg.paste(img, mask=img.split()[3])
        bg = bg.convert("RGB")
        buf = io.BytesIO()
        bg.save(buf, format="JPEG", quality=88)
        return buf.getvalue()
    except Exception:
        return raw_bytes


def save_photos(image_b64, mime_type, existing_filename=None):
    """Save original + bg-removed versions. Returns filename."""
    import shutil
    if "," in image_b64:
        image_b64 = image_b64.split(",", 1)[1]
    raw = base64.b64decode(image_b64)
    # normalise to JPEG
    img = Image.open(io.BytesIO(raw)).convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    raw_jpg = buf.getvalue()

    filename = existing_filename or f"{uuid.uuid4().hex[:8]}.jpg"
    orig_dir = os.path.join(os.path.dirname(PHOTOS_DIR), "photos_original")
    proc_dir = os.path.join(os.path.dirname(PHOTOS_DIR), "photos_processed")
    os.makedirs(orig_dir, exist_ok=True)
    os.makedirs(proc_dir, exist_ok=True)

    with open(os.path.join(orig_dir, filename), "wb") as f:
        f.write(raw_jpg)

    processed = remove_background(raw_jpg)
    with open(os.path.join(proc_dir, filename), "wb") as f:
        f.write(processed)

    # Active photo defaults to processed
    with open(os.path.join(PHOTOS_DIR, filename), "wb") as f:
        f.write(processed)

    return filename


def shrink_image(image_b64, mime_type='image/jpeg', max_px=1024, quality=82):
    """Resize image to max_px on longest side before sending to API."""
    try:
        raw = base64.b64decode(image_b64)
        img = Image.open(io.BytesIO(raw)).convert('RGB')
        w, h = img.size
        if max(w, h) > max_px:
            scale = max_px / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=quality)
        return base64.b64encode(buf.getvalue()).decode(), 'image/jpeg'
    except Exception:
        return image_b64, mime_type


ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_MODEL   = "claude-haiku-4-5-20251001"


def get_anthropic_key():
    try:
        with open(OPENCLAW_CONFIG) as f:
            return json.load(f)["env"]["ANTHROPIC_API_KEY"]
    except Exception:
        return os.environ.get("ANTHROPIC_API_KEY", "")


def identify_badge(image_b64, mime_type="image/jpeg"):
    image_b64, mime_type = shrink_image(image_b64, mime_type, max_px=1024)

    prompt = """You are a world-class expert on car badges, emblems, hood ornaments, wheel caps, and automotive logos.

A kid is showing you a photo of a car badge they found. Identify what you actually see — look carefully at the real shape, text, logo, colors, and chrome details in the photo. Make your best guess even if the photo is blurry or partial.

Reply with ONLY a JSON object using these exact keys (no markdown, no code blocks, no explanation):
{"make": "<actual car brand you see>", "model": "<specific model or All Models>", "years": "<year range>", "badge_name": "<what this badge is called>", "description": "<exciting kid-friendly description of what you see>", "fun_fact": "<cool fact about this car or badge>", "rarity": "<common|uncommon|rare|very_rare>", "value_estimate": "<dollar range>", "confidence": "<high|medium|low>"}

Rarity: "common" = mass-produced millions, "uncommon" = specific model/older variants, "rare" = pre-1970s/limited/European exotics, "very_rare" = pre-war/prototype/ultra-limited.
Keep description and fun_fact exciting and kid-friendly (8-12 year old audience).
If genuinely not a car badge, set make to "Unknown"."""

    payload = json.dumps({
        "model": ANTHROPIC_MODEL,
        "max_tokens": 500,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {
                    "type": "base64",
                    "media_type": mime_type,
                    "data": image_b64
                }},
                {"type": "text", "text": prompt}
            ]
        }]
    }).encode()

    req = urllib.request.Request(
        ANTHROPIC_API_URL,
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": get_anthropic_key(),
            "anthropic-version": "2023-06-01"
        }
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            resp = json.loads(r.read())
            content = resp["content"][0]["text"].strip()
            if "```" in content:
                content = content.split("```")[1]
                if content.startswith("json"):
                    content = content[4:]
                content = content.strip()
            start = content.find("{")
            end   = content.rfind("}") + 1
            if start >= 0 and end > start:
                content = content[start:end]
            return json.loads(content)
    except Exception as e:
        return {"error": str(e), "make": "Unknown", "badge_name": "Unidentified Badge",
                "description": "Couldn't identify this one — try a clearer photo!", "confidence": "low"}


def generate_ideas(owned_makes):
    owned_str = ", ".join(sorted(set(m for m in owned_makes if m))) or "none yet"

    prompt = f"""You are a fun, encouraging mentor helping a kid grow his car badge/emblem collection.

He already has badges from these makes: {owned_str}.

He specifically wants RARE, HARD-TO-GET EXOTIC badges — Ferrari, Lamborghini, Porsche, Bentley, Rolls-Royce, Aston Martin, Bugatti, Maserati, McLaren — the kind you basically never find at a regular self-service junkyard.

Give him 6 creative, realistic ideas (assume a parent/adult can help with driving, asking, etc.) for how to actually find or get one of these rare badges. Mix easy wins with bigger swings. Think broadly: car shows and cruise nights, exotic dealership service departments, owners' clubs (Ferrari Club of America, Porsche Club of America, etc.), collision-repair/specialty shops, online collector marketplaces and swap groups, estate sales, Cars & Coffee events, trading with other badge collectors.

Reply with ONLY a JSON object, no markdown, no code fences:
{{"ideas": [{{"title": "<short punchy title>", "difficulty": "<easy|medium|hard>", "description": "<2-3 exciting kid-friendly sentences on what to actually do>", "tip": "<one concrete practical tip, e.g. what to say or where to look>"}}]}}"""

    payload = json.dumps({
        "model": ANTHROPIC_MODEL,
        "max_tokens": 900,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()

    req = urllib.request.Request(
        ANTHROPIC_API_URL,
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": get_anthropic_key(),
            "anthropic-version": "2023-06-01"
        }
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            resp = json.loads(r.read())
            content = resp["content"][0]["text"].strip()
            if "```" in content:
                content = content.split("```")[1]
                if content.startswith("json"):
                    content = content[4:]
                content = content.strip()
            start = content.find("{")
            end   = content.rfind("}") + 1
            if start >= 0 and end > start:
                content = content[start:end]
            return json.loads(content)
    except Exception as e:
        return {"error": str(e), "ideas": []}


RARITY_ORDER = ['very_rare', 'rare', 'uncommon', 'common', '']

FILLS = {
    'very_rare': PatternFill('solid', fgColor='FFD700'),
    'rare':      PatternFill('solid', fgColor='FFB347'),
    'uncommon':  PatternFill('solid', fgColor='B0D4E8'),
    'common':    PatternFill('solid', fgColor='F0F0F0'),
    '':          PatternFill('solid', fgColor='F8F8F8'),
}
RARITY_LABEL = {
    'very_rare': '🔥 Very Rare',
    'rare':      '⭐⭐⭐ Rare',
    'uncommon':  '⭐⭐ Uncommon',
    'common':    '⭐ Common',
}
NAVY_FILL  = PatternFill('solid', fgColor='1D3557')
WHITE_BOLD = Font(bold=True, color='FFFFFF', size=11)
THIN       = Border(bottom=Side(style='thin', color='DDDDDD'))


def build_excel(badges):
    wb = Workbook()

    # ── Sheet 1: Full Collection ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'Badge Collection'

    headers = ['#', 'Make', 'Model', 'Badge Name', 'Years', 'Rarity', 'Est. Value', 'Description', 'Fun Fact', 'Date Added']
    widths  = [4,   16,     20,      36,            14,      16,       12,            60,             60,          14]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = WHITE_BOLD
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

    sorted_badges = sorted(badges, key=lambda b: RARITY_ORDER.index(b.get('rarity', '')) if b.get('rarity', '') in RARITY_ORDER else 99)

    for ri, b in enumerate(sorted_badges, 2):
        rarity = b.get('rarity', '')
        fill   = FILLS.get(rarity, FILLS[''])
        added  = b.get('added', '')[:10] if b.get('added') else ''
        row_data = [
            ri - 1,
            b.get('make', ''),
            b.get('model', ''),
            b.get('badge_name', ''),
            b.get('years', ''),
            RARITY_LABEL.get(rarity, rarity),
            b.get('value_estimate', ''),
            b.get('description', ''),
            b.get('fun_fact', ''),
            added,
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(ri, ci, val)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = THIN
        ws.row_dimensions[ri].height = 60

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # ── Sheet 2: Stats ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Stats')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 10

    ws2['A1'] = "Matthew's Badge Collection"
    ws2['A1'].font = Font(bold=True, size=16, color='1D3557')
    ws2.merge_cells('A1:B1')
    ws2.row_dimensions[1].height = 28

    ws2['A2'] = f"Generated {datetime.now().strftime('%B %d, %Y')}"
    ws2['A2'].font = Font(italic=True, color='888888', size=10)
    ws2.merge_cells('A2:B2')

    counts = {}
    for b in badges:
        r = b.get('rarity', 'unknown')
        counts[r] = counts.get(r, 0) + 1
    makes = len(set(b.get('make', '') for b in badges if b.get('make')))

    stats = [
        ('Total Badges', len(badges)),
        ('Different Makes', makes),
        ('Very Rare 🔥', counts.get('very_rare', 0)),
        ('Rare ⭐⭐⭐', counts.get('rare', 0)),
        ('Uncommon ⭐⭐', counts.get('uncommon', 0)),
        ('Common ⭐', counts.get('common', 0)),
    ]

    for ri, (label, val) in enumerate(stats, 4):
        ws2.cell(ri, 1, label).font = Font(bold=True, color='1D3557')
        ws2.cell(ri, 2, val).alignment = Alignment(horizontal='center')
        ws2.cell(ri, 2).font = Font(bold=True, size=13, color='1D3557')
        for ci in range(1, 3):
            ws2.cell(ri, ci).border = THIN

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def load_collection():
    if os.path.exists(COLLECTION_FILE):
        with open(COLLECTION_FILE) as f:
            data = json.load(f)
            # Support both old format {"badges":[]} and new format with metadata
            if isinstance(data, list):
                return {"badges": data}
            return data
    return {"badges": []}


def save_collection(collection, badge_name=None):
    with open(COLLECTION_FILE, "w") as f:
        json.dump(collection, f, indent=2)
    # Push to GitHub in background so it doesn't slow down the response
    script = os.path.join(os.path.dirname(BASE_DIR), "push-badge.sh")
    if os.path.exists(script):
        label = badge_name or "badge"
        subprocess.Popen(["/bin/bash", script, label],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


MIME_TYPES = {
    ".html": "text/html",
    ".css": "text/css",
    ".js": "application/javascript",
    ".json": "application/json",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".svg": "image/svg+xml",
    ".ico": "image/x-icon",
    ".webp": "image/webp",
}


class BadgeHandler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass  # Quiet logging

    def send_json(self, data, status=200):
        body = json.dumps(data).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length) if length else b""

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/" or path == "/index.html":
            self.serve_file(os.path.join(BASE_DIR, "index.html"))
        elif path == "/admin" or path == "/admin.html":
            self.serve_file(os.path.join(BASE_DIR, "admin.html"))
        elif path == "/gallery" or path == "/gallery.html":
            self.serve_file(os.path.join(BASE_DIR, "gallery.html"))
        elif path == "/api/collection":
            self.send_json(load_collection())
        elif path == "/api/export/excel":
            collection = load_collection()
            xlsx_bytes = build_excel(collection.get("badges", []))
            filename = f"MatthewBadges_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", len(xlsx_bytes))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            return
        elif path.startswith("/photos/"):
            photo_name = path[8:]
            self.serve_file(os.path.join(PHOTOS_DIR, photo_name))
        elif path.startswith("/photos_original/"):
            photo_name = path[17:]
            self.serve_file(os.path.join(os.path.dirname(PHOTOS_DIR), "photos_original", photo_name))
        elif path.startswith("/photos_processed/"):
            photo_name = path[18:]
            self.serve_file(os.path.join(os.path.dirname(PHOTOS_DIR), "photos_processed", photo_name))
        else:
            file_path = os.path.join(BASE_DIR, path.lstrip("/"))
            if os.path.isfile(file_path):
                self.serve_file(file_path)
            else:
                self.send_json({"error": "Not found"}, 404)

    def serve_file(self, file_path):
        if not os.path.isfile(file_path):
            self.send_response(404)
            self.end_headers()
            return
        ext = os.path.splitext(file_path)[1].lower()
        mime = MIME_TYPES.get(ext, "application/octet-stream")
        with open(file_path, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", len(data))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/ideas":
            collection = load_collection()
            owned_makes = [b.get("make", "") for b in collection.get("badges", [])]
            result = generate_ideas(owned_makes)
            self.send_json(result)

        elif path == "/api/identify":
            body = json.loads(self.read_body())
            image_b64 = body.get("image", "")
            mime_type = body.get("mimeType", "image/jpeg")
            # Strip data URL prefix if present
            if "," in image_b64:
                image_b64 = image_b64.split(",", 1)[1]
            result = identify_badge(image_b64, mime_type)
            self.send_json(result)

        elif path == "/api/add":
            body = json.loads(self.read_body())
            badge_info = body.get("badge", {})
            image_b64 = body.get("image", "")
            mime_type = body.get("mimeType", "image/jpeg")

            photo_filename = None
            if image_b64:
                photo_filename = save_photos(image_b64, mime_type)

            collection = load_collection()
            entry = {
                "id": uuid.uuid4().hex[:8],
                "added": datetime.now().isoformat(),
                "photo": f"/photos/{photo_filename}" if photo_filename else None,
                **badge_info
            }
            collection["badges"].append(entry)
            save_collection(collection, badge_name=badge_info.get("badge_name"))
            self.send_json({"ok": True, "entry": entry})

        elif path.startswith("/api/badge/") and path.endswith("/update"):
            # Update badge identification in place, keeping same photo
            badge_id = path[11:-7]
            body = json.loads(self.read_body())
            new_info = body.get("badge", {})
            collection = load_collection()
            for i, b in enumerate(collection["badges"]):
                if b.get("id") == badge_id:
                    # Preserve id, added, photo — replace everything else
                    preserved = {k: b[k] for k in ("id", "added", "photo") if k in b}
                    collection["badges"][i] = {**preserved, **new_info}
                    save_collection(collection)
                    self.send_json({"ok": True, "entry": collection["badges"][i]})
                    return
            self.send_json({"error": "Not found"}, 404)

        elif path == "/api/reidentify":
            # Re-run AI identification on a badge's saved photo
            body = json.loads(self.read_body())
            badge_id = body.get("id")
            collection = load_collection()
            badge = next((b for b in collection["badges"] if b.get("id") == badge_id), None)
            if not badge or not badge.get("photo"):
                self.send_json({"error": "Badge or photo not found"}, 404)
                return
            photo_path = os.path.join(BASE_DIR, badge["photo"].lstrip("/"))
            if not os.path.isfile(photo_path):
                self.send_json({"error": "Photo file missing"}, 404)
                return
            with open(photo_path, "rb") as f:
                image_b64 = base64.b64encode(f.read()).decode()
            result = identify_badge(image_b64, "image/jpeg")
            self.send_json(result)

        elif path.startswith("/api/badge/") and path.endswith("/rotate"):
            import shutil
            badge_id = path[11:-7]
            body = json.loads(self.read_body())
            degrees = int(body.get("degrees", 90))  # 90 or -90
            collection = load_collection()
            badge = next((b for b in collection["badges"] if b.get("id") == badge_id), None)
            if not badge or not badge.get("photo"):
                self.send_json({"error": "Badge not found"}, 404)
                return
            filename = os.path.basename(badge["photo"])
            # Rotate all three copies (photos, photos_original, photos_processed)
            for folder in [PHOTOS_DIR,
                           os.path.join(os.path.dirname(PHOTOS_DIR), "photos_original"),
                           os.path.join(os.path.dirname(PHOTOS_DIR), "photos_processed")]:
                fp = os.path.join(folder, filename)
                if os.path.isfile(fp):
                    img = Image.open(fp).convert("RGB")
                    img = img.rotate(-degrees, expand=True)  # PIL rotates CCW, flip sign
                    img.save(fp, format="JPEG", quality=90)
            save_collection(collection, badge_name=f"rotate {badge_id}")
            self.send_json({"ok": True})

        elif path.startswith("/api/badge/") and path.endswith("/retake"):
            badge_id = path[11:-7]
            body = json.loads(self.read_body())
            image_b64 = body.get("image", "")
            mime_type = body.get("mimeType", "image/jpeg")
            if not image_b64:
                self.send_json({"error": "No image provided"}, 400)
                return
            collection = load_collection()
            badge = next((b for b in collection["badges"] if b.get("id") == badge_id), None)
            if not badge:
                self.send_json({"error": "Badge not found"}, 404)
                return
            # Reuse existing filename if possible so paths stay consistent
            existing_fn = os.path.basename(badge.get("photo", "")) or None
            filename = save_photos(image_b64, mime_type, existing_filename=existing_fn)
            badge["photo"] = f"/photos/{filename}"
            save_collection(collection, badge_name=f"retake {badge_id}")
            self.send_json({"ok": True, "photo": badge["photo"]})

        elif path.startswith("/api/badge/") and path.endswith("/use-photo"):
            # Swap the active photo for a badge: {"version": "original" | "processed"}
            import shutil
            badge_id = path[11:-10]
            body     = json.loads(self.read_body())
            version  = body.get("version", "processed")
            photos_dir   = PHOTOS_DIR
            originals_dir = os.path.join(os.path.dirname(photos_dir), "photos_original")
            collection = load_collection()
            badge = next((b for b in collection["badges"] if b.get("id") == badge_id), None)
            if not badge:
                self.send_json({"error": "Badge not found"}, 404)
                return
            filename = os.path.basename(badge.get("photo", ""))
            if not filename:
                self.send_json({"error": "No photo on badge"}, 404)
                return
            src = os.path.join(originals_dir if version == "original" else
                               os.path.join(os.path.dirname(photos_dir), "photos_processed"),
                               filename)
            dst = os.path.join(photos_dir, filename)
            if not os.path.isfile(src):
                self.send_json({"error": f"{version} photo not found"}, 404)
                return
            shutil.copy2(src, dst)
            save_collection(collection, badge_name=f"update photo {badge_id}")
            self.send_json({"ok": True, "version": version})

        else:
            self.send_json({"error": "Not found"}, 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/badge/"):
            badge_id = path[11:]
            collection = load_collection()
            before = len(collection["badges"])
            collection["badges"] = [b for b in collection["badges"] if b.get("id") != badge_id]
            removed = before - len(collection["badges"])
            save_collection(collection, badge_name="remove badge" if removed else None)
            self.send_json({"ok": True, "removed": removed})
        else:
            self.send_json({"error": "Not found"}, 404)


class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True

if __name__ == "__main__":
    server = ThreadedHTTPServer(("0.0.0.0", PORT), BadgeHandler)
    print(f"🏆 Badge Hunter server running at http://localhost:{PORT}")
    print(f"   Photos stored in: {PHOTOS_DIR}")
    print(f"   Collection file:  {COLLECTION_FILE}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
